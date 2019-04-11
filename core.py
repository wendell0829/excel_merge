'''
    核心代码，包括读取excel内容至dict， 合并dict， 将dict写入excel三部分
'''


from openpyxl import load_workbook, Workbook
import os


def load_to_dict(filename: str):
    '''
    将excel文件读取到一个dict中，具体形式如下：
    ｛
        ‘sheet name’：
            ｛
                列标题： [ 列内容 ]
            ｝
    ｝
    :param filename:  要打开的excel文件路径， str类型
    :return:  转换后生成的dict
    '''
    workbook_dict = {}
    workbook = load_workbook(filename)
    for sheet in workbook.worksheets:
        sheet_dict = {}
        for column in sheet.columns:
            # 遍历每一列， 第一行作为列标题， 第二行以后作为列内容
            sheet_dict[column[0].value] = [i.value for i in column[1:]]
        workbook_dict[sheet.title] = sheet_dict

    return workbook_dict


def merge_workbook_dict(workbook_dict1: dict, workbook_dict2: dict):
    '''
    将两个dict合并， 只有相同sheet名称下相同列标题的内容会被合并， 其余被舍弃
    :param workbook_dict1: 要合并的dict
    :param workbook_dict2: 要合并的dict
    :return: 合并后的dict
    '''
    end_dict = workbook_dict1.copy()
    for sheet in workbook_dict2.keys():
        if sheet in end_dict.keys():
            for title in workbook_dict2[sheet]:
                if title in end_dict[sheet].keys():
                    end_dict[sheet][title].extend(workbook_dict2[sheet][title])
    return end_dict


def write_dict_to_excel(workbook_dict: dict, filename: str):
    '''
    将一个dict写入excel文件， dict的形式必须严格符合上述函数规定的格式， 请不要随意调用
    :param workbook_dict: 符合格式的dict
    :param filename: 生成的文件名
    :return: 无， 文件保存完毕后打印成功信息
    '''
    filename = 'files/' + filename + '.xlsx'
    if os.path.exists(filename):
        print('文件名已存在！')
        return
    workbook = Workbook() # 实例化一个workbook对象
    for sheet_name in workbook_dict.keys():
        workbook.create_sheet(title=sheet_name) # 遍历字典，取出所有sheet名称， 并新建sheet页
        workbook[sheet_name].append(list(workbook_dict[sheet_name].keys())) # 利用sheet的append方法将列标题直接写入第一行
        for title in workbook_dict[sheet_name].keys():
            for title_cell in workbook[sheet_name][1]:
                if title == title_cell.value:
                    for r in range(len(workbook_dict[sheet_name][title])):
                        # 遍历字典中的列标题与sheet中第一行的列标题， 有相同的则将列内容写入
                        workbook[sheet_name].cell(row=r+2, column=title_cell.col_idx).value = workbook_dict[sheet_name][title][r]

    workbook.remove(workbook.worksheets[0]) # 生成workbook对象时会自动生成一个空白sheet， 这里删掉
    workbook.save(filename=filename)
    print('文件生成完毕，请在files文件夹中查看！')

